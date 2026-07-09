###################################################
# Project:     IOManager
# Author:      Peter Chrapchynski
# Date:        2026Jun23
# History:     2026Jun23 - Initial creation
#              2026Jul08 - Add tests for Skip column
###################################################

import io
import openpyxl
import pytest
from pathlib import Path
from tests.conftest import IO_HEADERS, make_excel_bytes
from app.services.etl.excel_reader import ExcelReader, build_log_messages


def _write_excel(tmp_path: Path, rows: list[dict]) -> Path:
    path = tmp_path / "io_index.xlsx"
    path.write_bytes(make_excel_bytes(rows))
    return path


@pytest.fixture
def reader():
    return ExcelReader()


class TestRead:
    def test_reads_single_row(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Number": 1, "Tag Name": "LSL-001", "Description": "Low level",
             "Template": "DI"}
        ])
        rows = reader.read(path)
        assert len(rows) == 1

    def test_tag_name_preserved(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LAL-001", "Template": "DI"}
        ])
        rows = reader.read(path)
        assert rows[0].tag_name == "LAL-001"

    def test_description_preserved(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Description": "Tank low level", "Template": "DI"}
        ])
        rows = reader.read(path)
        assert rows[0].description == "Tank low level"

    def test_skips_row_without_tag_name(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": None, "Template": "DI"},
            {"Tag Name": "LAL-001", "Template": "DI"},
        ])
        rows = reader.read(path)
        assert len(rows) == 1
        assert rows[0].tag_name == "LAL-001"

    def test_skips_row_without_template(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": None},
            {"Tag Name": "LAL-001", "Template": "DI"},
        ])
        rows = reader.read(path)
        assert len(rows) == 1

    def test_skips_row_with_skip_equals_1(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI", "Skip": 1},
            {"Tag Name": "LAL-001", "Template": "DI"},
        ])
        rows = reader.read(path)
        assert len(rows) == 1
        assert rows[0].tag_name == "LAL-001"

    def test_includes_row_with_skip_equals_0(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI", "Skip": 0},
        ])
        rows = reader.read(path)
        assert len(rows) == 1

    def test_includes_row_with_skip_absent(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI"},
        ])
        rows = reader.read(path)
        assert len(rows) == 1

    def test_module_channel_parsed_as_int(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "DIG-001", "Template": "DI", "Module": 3, "Module Channel": 7}
        ])
        rows = reader.read(path)
        assert rows[0].module == 3
        assert rows[0].module_channel == 7

    def test_failsafe_parsed_as_bool(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI", "Failsafe": 1}
        ])
        rows = reader.read(path)
        assert rows[0].failsafe is True

    def test_is_alarm_parsed_as_bool(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI", "isAlm": 1,
             "AlmMsg": "Low level"}
        ])
        rows = reader.read(path)
        assert rows[0].is_alarm is True

    def test_alarm_condition_uppercased(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI", "isAlm": 1,
             "AlmCondition": "pos"}
        ])
        rows = reader.read(path)
        assert rows[0].alarm_condition == "POS"

    def test_multiple_rows_in_order(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI"},
            {"Tag Name": "LAL-001", "Template": "DI"},
            {"Tag Name": "XY-001",  "Template": "DO"},
        ])
        rows = reader.read(path)
        assert [r.tag_name for r in rows] == ["LSL-001", "LAL-001", "XY-001"]

    def test_row_number_is_excel_row_index(self, reader, tmp_path):
        # First data row is Excel row 2 (header is row 1)
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI"},
        ])
        rows = reader.read(path)
        assert rows[0].number == 2

    def test_missing_sheet_raises(self, reader, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Wrong Sheet"
        path = tmp_path / "bad.xlsx"
        buf = io.BytesIO()
        wb.save(buf)
        path.write_bytes(buf.getvalue())
        with pytest.raises(ValueError, match="IO Dist"):
            reader.read(path)


class TestWriteLog:
    def test_writes_success_message(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "DI"}
        ])
        reader.write_log(path, {2: "Processed at 01/01/2026 12:00:00"}, set())
        wb = openpyxl.load_workbook(path)
        ws = wb["IO Dist"]
        log_col = next(
            i + 1 for i, c in enumerate(ws[1]) if str(c.value).strip() == "Log"
        )
        assert ws.cell(row=2, column=log_col).value == "Processed at 01/01/2026 12:00:00"

    def test_error_row_gets_red_fill(self, reader, tmp_path):
        path = _write_excel(tmp_path, [
            {"Tag Name": "LSL-001", "Template": "BADTEMPLATE"}
        ])
        reader.write_log(path, {2: "Template: (BADTEMPLATE) not found"}, {2})
        wb = openpyxl.load_workbook(path)
        ws = wb["IO Dist"]
        log_col = next(
            i + 1 for i, c in enumerate(ws[1]) if str(c.value).strip() == "Log"
        )
        fill = ws.cell(row=2, column=log_col).fill
        # openpyxl stores ARGB; fully-opaque red = FFFF0000
        assert fill.start_color.rgb.upper().endswith("FF0000")


class TestBuildLogMessages:
    def test_success_row_gets_processed_message(self, tmp_path):
        from app.models.address_map import AddressMap
        from app.models.config import AppConfig
        from app.services.etl.rule_engine import RuleEngine
        from app.models.io_row import IoIndexRow

        config_json = Path(__file__).parent / "fixtures" / "app.config.json"
        config = AppConfig.model_validate_json(config_json.read_text())
        am = AddressMap()
        engine = RuleEngine(config=config, address_map=am)

        rows = [IoIndexRow(number=2, tag_name="LSL-001", description="Level", template="DI")]
        result = engine.process(rows)

        logs, error_rows = build_log_messages(result, rows)
        assert 2 in logs
        assert "Processed at" in logs[2]
        assert 2 not in error_rows

    def test_error_row_gets_error_message(self, tmp_path):
        from app.models.address_map import AddressMap
        from app.models.config import AppConfig
        from app.services.etl.rule_engine import RuleEngine
        from app.models.io_row import IoIndexRow

        config_json = Path(__file__).parent / "fixtures" / "app.config.json"
        config = AppConfig.model_validate_json(config_json.read_text())
        am = AddressMap()
        engine = RuleEngine(config=config, address_map=am)

        rows = [IoIndexRow(number=3, tag_name="XX-001", template="UNKNOWN")]
        result = engine.process(rows)

        logs, error_rows = build_log_messages(result, rows)
        assert 3 in logs
        assert "UNKNOWN" in logs[3]
        assert 3 in error_rows
