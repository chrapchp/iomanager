###################################################
# Project:     IOManager
# Author:      Peter Chrapchynski
# Date:        2026Jun23
# History:     2026Jun23 - Initial creation
#              2026Jun24 - Sheet lookup now case-insensitive and whitespace-tolerant
###################################################

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

from app.models.io_row import IoIndexRow

_SHEET_NAME = "IO Dist"

# Normalised header text → IoIndexRow field name
_HEADER_MAP: dict[str, str] = {
    "number": "number",
    "tag name": "tag_name",
    "description": "description",
    "i/o type": "io_type",
    "part number": "part_number",
    "module": "module",
    "module channel": "module_channel",
    "connector": "connector",
    "connector channel": "connector_channel",
    "signal": "signal",
    "phase": "phase",
    "note": "note",
    "template": "template",
    "failsafe": "failsafe",
    "haspresentation": "has_presentation",
    "presentation": "presentation",
    "units": "units",
    "inputmax": "input_max",
    "inputmin": "input_min",
    "isalm": "is_alarm",
    "almcondition": "alarm_condition",
    "almmsg": "alarm_message",
}

_RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
_INT_FIELDS = {"module", "module_channel", "connector", "connector_channel", "phase", "number"}
_BOOL_FIELDS = {"failsafe", "has_presentation", "is_alarm"}


class ExcelReader:
    """Reads and writes the IO Dist tab of the engineer's Excel file."""

    def read(self, file_path: Path) -> list[IoIndexRow]:
        """Parse the IO Dist tab and return a list of IoIndexRow objects."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        target = _SHEET_NAME.strip().lower()
        actual = next((s for s in wb.sheetnames if s.strip().lower() == target), None)
        if actual is None:
            raise ValueError(
                f"Sheet '{_SHEET_NAME}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[actual]

        col_index = self._parse_header(ws)
        rows: list[IoIndexRow] = []

        for excel_row, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            raw: dict[str, Any] = {
                field: row_cells[col_idx].value
                for field, col_idx in col_index.items()
            }

            # skip rows with no tag name
            if not raw.get("tag_name"):
                continue

            # skip rows with no template assigned
            if not raw.get("template"):
                continue

            row = self._coerce(excel_row, raw)
            if row is not None:
                rows.append(row)

        return rows

    def write_log(
        self,
        file_path: Path,
        logs: dict[int, str],
        error_rows: set[int],
    ) -> None:
        """
        Write processing status back to the Log column.

        logs: {excel_row_number: message_string}
        error_rows: set of excel row numbers to highlight red
        """
        wb = openpyxl.load_workbook(file_path)
        target = _SHEET_NAME.strip().lower()
        actual = next((s for s in wb.sheetnames if s.strip().lower() == target), None)
        if actual is None:
            return

        ws = wb[actual]
        col_index = self._parse_header(ws)

        if "log" not in col_index:
            return

        log_col_idx = col_index["log"] + 1  # openpyxl uses 1-based column

        for excel_row, message in logs.items():
            cell = ws.cell(row=excel_row, column=log_col_idx)
            cell.value = message
            if excel_row in error_rows:
                cell.fill = _RED_FILL

        wb.save(file_path)

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_header(ws: Any) -> dict[str, int]:
        """Returns {field_name: column_index (0-based)} for known columns."""
        col_index: dict[str, int] = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value is None:
                continue
            normalised = str(cell.value).strip().lower()
            field = _HEADER_MAP.get(normalised)
            if field:
                col_index[field] = idx
            elif normalised == "log":
                col_index["log"] = idx
        return col_index

    @staticmethod
    def _coerce(excel_row: int, raw: dict[str, Any]) -> IoIndexRow | None:
        """Coerce raw cell values to IoIndexRow-compatible types."""
        data: dict[str, Any] = {"number": excel_row}

        for field, value in raw.items():
            if field == "number":
                continue  # excel_row is authoritative; cell value may be blank
            if field in _INT_FIELDS:
                if value is not None:
                    try:
                        data[field] = int(value)
                    except (ValueError, TypeError):
                        data[field] = None
                else:
                    data[field] = None
            elif field in _BOOL_FIELDS:
                data[field] = bool(value) if value else False
            elif field in ("input_min", "input_max"):
                data[field] = str(value) if value is not None else ""
            elif field == "alarm_condition":
                data[field] = str(value).strip().upper() if value else None
            else:
                data[field] = str(value).strip() if value is not None else ""

        try:
            return IoIndexRow(**data)
        except Exception:
            return None


def build_log_messages(
    result: Any,  # GenerationResult
    rows: list[IoIndexRow],
    excel_base_row: int = 2,
) -> tuple[dict[int, str], set[int]]:
    """
    Build the logs dict for Excel write-back from a GenerationResult.

    Returns (logs, error_rows) where logs maps excel row → message.
    """
    error_by_row = {e.row_number: e for e in result.errors}
    logs: dict[int, str] = {}
    error_rows: set[int] = set()

    for row in rows:
        excel_row = row.number  # stored as Excel row number during read
        if excel_row in error_by_row:
            err = error_by_row[excel_row]
            logs[excel_row] = err.message
            error_rows.add(excel_row)
        else:
            logs[excel_row] = (
                "Processed at " + datetime.now().strftime("%m/%d/%Y %H:%M:%S")
            )

    return logs, error_rows
